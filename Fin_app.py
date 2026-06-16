"""
Personal Finance Dashboard — Streamlit Web Application
=======================================================
A comprehensive personal finance tracker with:
- Excel-based portable database (auto-initialization)
- Interactive income & expense input with session state protection
- KPI metrics dashboard with interactive Plotly charts
- Rule-based Perspective Analytics (50/30/20, 80% threshold, non-essential overspend)
- Annual Excel report export functionality
"""

import os
from datetime import datetime, date
from io import BytesIO
from copy import deepcopy
from typing import Optional

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import plotly.graph_objects as go
    import plotly.express as px
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    go = None
    px = None

# =============================================================================
# CONSTANTS & CONFIGURATION
# =============================================================================

DB_DIR = "database"
INCOME_FILE = os.path.join(DB_DIR, "income_monthly.xlsx")
EXPENSE_FILE = os.path.join(DB_DIR, "expenses_monthly.xlsx")

INCOME_COLUMNS = ["Date", "Source", "Amount", "Month", "Year"]
EXPENSE_COLUMNS = ["Date", "Category", "Amount", "Month", "Year"]

# Categories used for analytics classification (50/30/20 rule)
NON_ESSENTIAL_CATEGORIES = ["Hiburan", "Gaya Hidup"]
NEEDS_CATEGORIES = ["Makanan", "Transportasi", "Tagihan"]

MONTH_NAMES_ID = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember"
]


# =============================================================================
# CUSTOM DATA FRAME (lightweight replacement for pandas DataFrame)
# =============================================================================

class SimpleDataFrame:
    """A lightweight DataFrame replacement that works without numpy/pandas."""

    def __init__(self, columns: list[str], data: Optional[list[dict]] = None):
        self.columns = list(columns)
        self._data: list[dict] = []
        if data:
            for row in data:
                self._data.append({col: row.get(col, None) for col in self.columns})

    def __len__(self) -> int:
        return len(self._data)

    def __bool__(self) -> bool:
        return len(self._data) > 0

    def __iter__(self):
        return iter(self._data)

    def __getitem__(self, key):
        if isinstance(key, str):
            return [row[key] for row in self._data]
        if isinstance(key, int):
            return self._data[key]
        if isinstance(key, slice):
            new_df = SimpleDataFrame(self.columns)
            new_df._data = self._data[key]
            return new_df
        return self._data[key]

    def copy(self):
        new_df = SimpleDataFrame(self.columns)
        new_df._data = deepcopy(self._data)
        return new_df

    def tail(self, n: int = 5):
        new_df = SimpleDataFrame(self.columns)
        new_df._data = self._data[-n:] if len(self._data) >= n else self._data[:]
        return new_df

    def empty(self) -> bool:
        return len(self._data) == 0

    def append(self, row: dict):
        clean_row = {col: row.get(col, None) for col in self.columns}
        self._data.append(clean_row)

    def remove_at(self, index: int) -> bool:
        """Remove a row by index. Returns True if successful."""
        if 0 <= index < len(self._data):
            self._data.pop(index)
            return True
        return False

    def filter(self, conditions: dict) -> list[dict]:
        """Filter rows where all conditions match."""
        result = []
        for row in self._data:
            match = all(row.get(col) == val for col, val in conditions.items())
            if match:
                result.append(row)
        return result

    def filter_to_sdf(self, conditions: dict):
        """Filter and return a new SimpleDataFrame."""
        filtered = self.filter(conditions)
        new_sdf = SimpleDataFrame(self.columns)
        new_sdf._data = filtered
        return new_sdf

    def sum_column(self, column: str) -> float:
        return sum(row.get(column, 0) or 0 for row in self._data)

    def groupby_sum(self, column: str) -> dict:
        groups: dict = {}
        for row in self._data:
            key = row.get(column)
            amt = row.get("Amount", 0) or 0
            if key not in groups:
                groups[key] = 0.0
            groups[key] += amt
        return groups

    def unique(self, column: str) -> list:
        seen = set()
        result = []
        for row in self._data:
            val = row.get(column)
            if val not in seen:
                seen.add(val)
                result.append(val)
        return result

    def to_dicts(self) -> list[dict]:
        return deepcopy(self._data)

    def enumerate(self):
        """Yield (index, row_dict) pairs."""
        yield from enumerate(self._data)


# =============================================================================
# DATABASE INITIALIZATION
# =============================================================================

def initialize_database():
    """Auto-initialize the database folder and Excel files if they don't exist."""
    if not os.path.exists(DB_DIR):
        os.makedirs(DB_DIR)
        st.info(f"📁 Folder database '{DB_DIR}' telah dibuat.")

    if not os.path.exists(INCOME_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Income"
        for col_idx, col_name in enumerate(INCOME_COLUMNS, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        wb.save(INCOME_FILE)
        st.info(f"📄 File '{INCOME_FILE}' telah dibuat.")

    if not os.path.exists(EXPENSE_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"
        for col_idx, col_name in enumerate(EXPENSE_COLUMNS, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        wb.save(EXPENSE_FILE)
        st.info(f"📄 File '{EXPENSE_FILE}' telah dibuat.")


# =============================================================================
# DATA ACCESS LAYER
# =============================================================================

def _read_excel_to_sdf(filepath: str, columns: list[str]) -> SimpleDataFrame:
    """Read an Excel file into a SimpleDataFrame."""
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        sdf = SimpleDataFrame(columns)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                row_data = {}
                for idx, col_name in enumerate(columns):
                    val = row[idx] if idx < len(row) else None
                    if hasattr(val, 'strftime'):
                        val = val.date() if hasattr(val, 'date') else str(val)
                    row_data[col_name] = val
                sdf.append(row_data)
        wb.close()
        return sdf
    except (FileNotFoundError, Exception):
        return SimpleDataFrame(columns)


def _write_sdf_to_excel(filepath: str, sdf: SimpleDataFrame):
    """Write a SimpleDataFrame to an Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for col_idx, col_name in enumerate(sdf.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(sdf._data, 2):
        for col_idx, col_name in enumerate(sdf.columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))
    wb.save(filepath)
    wb.close()


def load_income_data() -> SimpleDataFrame:
    return _read_excel_to_sdf(INCOME_FILE, INCOME_COLUMNS)


def save_income_data(sdf: SimpleDataFrame):
    _write_sdf_to_excel(INCOME_FILE, sdf)


def load_expense_data() -> SimpleDataFrame:
    return _read_excel_to_sdf(EXPENSE_FILE, EXPENSE_COLUMNS)


def save_expense_data(sdf: SimpleDataFrame):
    _write_sdf_to_excel(EXPENSE_FILE, sdf)


# =============================================================================
# SESSION STATE MANAGEMENT
# =============================================================================

def init_session_state():
    if "income_submitted" not in st.session_state:
        st.session_state.income_submitted = False
    if "expense_submitted" not in st.session_state:
        st.session_state.expense_submitted = False
    if "income_date" not in st.session_state:
        st.session_state.income_date = date.today()
    if "income_source" not in st.session_state:
        st.session_state.income_source = ""
    if "income_amount" not in st.session_state:
        st.session_state.income_amount = 0.0
    if "expense_date" not in st.session_state:
        st.session_state.expense_date = date.today()
    if "expense_category" not in st.session_state:
        st.session_state.expense_category = ""
    if "expense_amount" not in st.session_state:
        st.session_state.expense_amount = 0.0


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def get_month_name(month_number: int) -> str:
    return MONTH_NAMES_ID[month_number - 1]


def format_rupiah(amount: float) -> str:
    if amount >= 0:
        return f"Rp {amount:,.0f}".replace(",", ".")
    else:
        return f"-Rp {abs(amount):,.0f}".replace(",", ".")


def format_date_for_display(d):
    if hasattr(d, 'strftime'):
        return d.strftime("%d-%m-%Y")
    return str(d)


# =============================================================================
# KPI CALCULATION
# =============================================================================

def calculate_kpis(sdf_income: SimpleDataFrame, sdf_expense: SimpleDataFrame,
                   selected_month: int, selected_year: int):
    month_name = get_month_name(selected_month)
    inc_month = sdf_income.filter({"Month": month_name, "Year": selected_year})
    exp_month = sdf_expense.filter({"Month": month_name, "Year": selected_year})

    total_income = sum(r.get("Amount", 0) or 0 for r in inc_month)
    total_expense = sum(r.get("Amount", 0) or 0 for r in exp_month)
    net_income = total_income - total_expense

    if total_income > 0:
        savings_rate = (net_income / total_income) * 100
    else:
        savings_rate = None

    return total_income, total_expense, net_income, savings_rate, inc_month, exp_month


# =============================================================================
# INTERACTIVE CHARTS
# =============================================================================

def create_comparison_chart(sdf_income: SimpleDataFrame, sdf_expense: SimpleDataFrame, year: int):
    income_by_month = []
    expense_by_month = []

    for m in range(1, 13):
        month_name = get_month_name(m)
        inc_val = sum(
            r.get("Amount", 0) or 0
            for r in sdf_income.filter({"Month": month_name, "Year": year})
        ) if sdf_income else 0.0

        exp_val = sum(
            r.get("Amount", 0) or 0
            for r in sdf_expense.filter({"Month": month_name, "Year": year})
        ) if sdf_expense else 0.0

        income_by_month.append(inc_val)
        expense_by_month.append(exp_val)

    if not PLOTLY_AVAILABLE:
        st.warning("Plotly tidak tersedia. Menggunakan tabel sebagai alternatif.")
        data_rows = []
        for m in range(12):
            data_rows.append({
                "Bulan": MONTH_NAMES_ID[m],
                "Pemasukan": format_rupiah(income_by_month[m]),
                "Pengeluaran": format_rupiah(expense_by_month[m])
            })
        st.table(data_rows)
        return None

    fig = go.Figure()
    fig.add_trace(go.Bar(x=MONTH_NAMES_ID, y=income_by_month, name="Pemasukan",
                         marker_color="#27AE60", hovertemplate="%{y:,.0f}<extra>Pemasukan</extra>"))
    fig.add_trace(go.Bar(x=MONTH_NAMES_ID, y=expense_by_month, name="Pengeluaran",
                         marker_color="#E74C3C", hovertemplate="%{y:,.0f}<extra>Pengeluaran</extra>"))
    fig.update_layout(
        title=f"Perbandingan Pemasukan vs Pengeluaran — {year}",
        xaxis_title="Bulan", yaxis_title="Jumlah (Rp)",
        barmode="group", hovermode="x unified", template="plotly_white",
        height=450,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    fig.update_yaxes(tickprefix="Rp ", tickformat=",.0f")
    return fig


# =============================================================================
# PERSPECTIVE ANALYTICS ENGINE
# =============================================================================

def run_perspective_analytics(sdf_income: SimpleDataFrame, sdf_expense: SimpleDataFrame,
                               selected_month: int, selected_year: int):
    total_income, total_expense, net_income, savings_rate, inc_month, exp_month = calculate_kpis(
        sdf_income, sdf_expense, selected_month, selected_year
    )

    analytics_results = []

    if total_income <= 0:
        analytics_results.append({
            "type": "info",
            "title": "Data Pemasukan Belum Tersedia",
            "message": "Analisis tidak dapat dilakukan karena belum ada data pemasukan untuk periode ini."
        })
        return analytics_results

    # RULE 1: 80% Threshold Warning
    expense_ratio = (total_expense / total_income) * 100 if total_income > 0 else 0

    if expense_ratio >= 80:
        excess_amount = total_expense - (0.8 * total_income)
        analytics_results.append({
            "type": "error",
            "title": f"⚠️ PERINGATAN: Pengeluaran Mencapai {expense_ratio:.1f}%",
            "message": (
                f"Total pengeluaran bulan ini mencapai **{expense_ratio:.1f}%** "
                f"dari total pemasukan (**{format_rupiah(total_income)}**). "
                f"Ini melampaui batas aman **80%**."
            ),
            "suggestion": (
                f"💡 **SARAN:** Anda perlu memotong pengeluaran setidaknya "
                f"sebesar **{format_rupiah(excess_amount)}** pada bulan depan "
                f"untuk kembali ke ambang aman. "
                f"Fokus pada pengurangan kategori non-esensial seperti "
                f"Hiburan dan Gaya Hidup."
            ),
            "metric": f"{expense_ratio:.1f}%",
            "threshold": "80%",
            "excess": format_rupiah(excess_amount)
        })

    # RULE 2: Non-Essential > 20% (check by keyword matching on category names)
    if exp_month:
        non_essential_total = sum(
            r.get("Amount", 0) or 0
            for r in exp_month
            if any(kw in (r.get("Category", "") or "").lower()
                   for kw in ["hiburan", "gaya hidup", "game", "voucher", "streaming", "hobi", "jalan-jalan"])
        )

        if non_essential_total > 0:
            non_essential_ratio = (non_essential_total / total_income) * 100
            if non_essential_ratio > 20:
                excess_non_essential = non_essential_total - (0.2 * total_income)
                analytics_results.append({
                    "type": "warning",
                    "title": f"⚠️ PERINGATAN: Pengeluaran Non-Esensial {non_essential_ratio:.1f}%",
                    "message": (
                        f"Pengeluaran kategori non-esensial (hiburan, gaya hidup, dll) "
                        f"mencapai **{non_essential_ratio:.1f}%** dari total pemasukan "
                        f"(**{format_rupiah(non_essential_total)}**). "
                        f"Ini melebihi batas wajar **20%** menurut alokasi 50/30/20."
                    ),
                    "suggestion": (
                        f"💡 **SARAN:** Kurangi pengeluaran non-esensial "
                        f"sebesar **{format_rupiah(excess_non_essential)}** pada "
                        f"bulan depan untuk mengikuti alokasi 50/30/20. "
                        f"Batasan ideal: maksimal **{format_rupiah(0.2 * total_income)}** "
                        f"per bulan untuk kategori ini."
                    ),
                    "metric": f"{non_essential_ratio:.1f}%",
                    "threshold": "20%",
                    "excess": format_rupiah(excess_non_essential)
                })

    # RULE 3: 50/30/20 Rule Evaluation
    if exp_month:
        needs_total = sum(
            r.get("Amount", 0) or 0
            for r in exp_month
            if any(kw in (r.get("Category", "") or "").lower()
                   for kw in ["makanan", "transportasi", "tagihan", "listrik", "air",
                              "internet", "bpjs", "sewa", "kos", "bahan bakar", "bensin"])
        )
        wants_total = sum(
            r.get("Amount", 0) or 0
            for r in exp_month
            if any(kw in (r.get("Category", "") or "").lower()
                   for kw in ["hiburan", "gaya hidup", "game", "voucher", "streaming", "hobi", "jalan-jalan"])
        )

        actual_savings = max(0, total_income - total_expense)
        needs_ratio = (needs_total / total_income) * 100 if total_income > 0 else 0
        wants_ratio_adjusted = (wants_total / total_income) * 100 if total_income > 0 else 0
        savings_ratio_val = (actual_savings / total_income) * 100 if total_income > 0 else 0

        analytics_results.append({
            "type": "info",
            "title": "📊 Evaluasi Alokasi 50/30/20",
            "message": (
                f"**Alokasi Aktual Anda:**\n\n"
                f"| Kategori | Ideal | Aktual | Status |\n"
                f"|----------|-------|--------|--------|\n"
                f"| 🏠 Kebutuhan (Needs) | 50% | **{needs_ratio:.1f}%** | "
                f"{'✅ Sesuai' if needs_ratio <= 55 else '⚠️ Perlu Diperhatikan'} |\n"
                f"| 🎮 Keinginan (Wants) | 30% | **{wants_ratio_adjusted:.1f}%** | "
                f"{'✅ Sesuai' if wants_ratio_adjusted <= 33 else '⚠️ Perlu Diperhatikan'} |\n"
                f"| 💰 Tabungan (Savings) | 20% | **{savings_ratio_val:.1f}%** | "
                f"{'✅ Baik' if savings_ratio_val >= 18 else '⚠️ Perlu Ditingkatkan'} |\n\n"
                f"**Detail Nominal:**\n"
                f"- Kebutuhan: {format_rupiah(needs_total)}\n"
                f"- Keinginan: {format_rupiah(wants_total)}\n"
                f"- Tabungan: {format_rupiah(actual_savings)}"
            ),
            "suggestion": None
        })

    return analytics_results


# =============================================================================
# EXPORT FUNCTIONALITY
# =============================================================================

def generate_annual_report(sdf_income: SimpleDataFrame, sdf_expense: SimpleDataFrame, year: int) -> BytesIO:
    output = BytesIO()
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

    def style_data_cell(cell, is_currency=False):
        cell.border = thin_border
        if is_currency:
            cell.number_format = 'Rp #,##0'

    # Sheet 1: Ringkasan Tahunan
    ws1 = wb.active
    ws1.title = "Ringkasan Tahunan"
    headers_1 = ["Bulan", "Total Pemasukan", "Total Pengeluaran", "Saldo Bersih", "Rasio Tabungan (%)"]
    for col_idx, h in enumerate(headers_1, 1):
        ws1.cell(row=1, column=col_idx, value=h)
    style_header(ws1, len(headers_1))

    for row_idx, m in enumerate(range(1, 13), 2):
        month_name = get_month_name(m)
        inc_val = sum(r.get("Amount", 0) or 0 for r in sdf_income.filter({"Month": month_name, "Year": year})) if sdf_income else 0.0
        exp_val = sum(r.get("Amount", 0) or 0 for r in sdf_expense.filter({"Month": month_name, "Year": year})) if sdf_expense else 0.0
        net = inc_val - exp_val
        saving_rate_val = ((net / inc_val) * 100) if inc_val > 0 else 0

        ws1.cell(row=row_idx, column=1, value=month_name)
        style_data_cell(ws1.cell(row=row_idx, column=1))
        style_data_cell(ws1.cell(row=row_idx, column=2, value=inc_val), is_currency=True)
        style_data_cell(ws1.cell(row=row_idx, column=3, value=exp_val), is_currency=True)
        style_data_cell(ws1.cell(row=row_idx, column=4, value=net), is_currency=True)
        c5 = ws1.cell(row=row_idx, column=5, value=saving_rate_val / 100)
        style_data_cell(c5)
        c5.number_format = '0.0%'

    for col, w in zip("ABCDE", [20, 20, 20, 20, 20]):
        ws1.column_dimensions[col].width = w

    # Sheet 2: Detail Pemasukan
    ws2 = wb.create_sheet("Detail Pemasukan")
    for col_idx, h in enumerate(INCOME_COLUMNS, 1):
        ws2.cell(row=1, column=col_idx, value=h)
    style_header(ws2, len(INCOME_COLUMNS))
    income_rows = sdf_income.filter({"Year": year}) if sdf_income else []
    for row_idx, r in enumerate(income_rows, 2):
        ws2.cell(row=row_idx, column=1, value=str(r.get("Date", "")))
        style_data_cell(ws2.cell(row=row_idx, column=1))
        ws2.cell(row=row_idx, column=2, value=r.get("Source", ""))
        style_data_cell(ws2.cell(row=row_idx, column=2))
        style_data_cell(ws2.cell(row=row_idx, column=3, value=r.get("Amount", 0)), is_currency=True)
        ws2.cell(row=row_idx, column=4, value=r.get("Month", ""))
        style_data_cell(ws2.cell(row=row_idx, column=4))
        ws2.cell(row=row_idx, column=5, value=r.get("Year", ""))
        style_data_cell(ws2.cell(row=row_idx, column=5))
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 15
    ws2.column_dimensions["E"].width = 10

    # Sheet 3: Detail Pengeluaran
    ws3 = wb.create_sheet("Detail Pengeluaran")
    for col_idx, h in enumerate(EXPENSE_COLUMNS, 1):
        ws3.cell(row=1, column=col_idx, value=h)
    style_header(ws3, len(EXPENSE_COLUMNS))
    expense_rows = sdf_expense.filter({"Year": year}) if sdf_expense else []
    for row_idx, r in enumerate(expense_rows, 2):
        ws3.cell(row=row_idx, column=1, value=str(r.get("Date", "")))
        style_data_cell(ws3.cell(row=row_idx, column=1))
        ws3.cell(row=row_idx, column=2, value=r.get("Category", ""))
        style_data_cell(ws3.cell(row=row_idx, column=2))
        style_data_cell(ws3.cell(row=row_idx, column=3, value=r.get("Amount", 0)), is_currency=True)
        ws3.cell(row=row_idx, column=4, value=r.get("Month", ""))
        style_data_cell(ws3.cell(row=row_idx, column=4))
        ws3.cell(row=row_idx, column=5, value=r.get("Year", ""))
        style_data_cell(ws3.cell(row=row_idx, column=5))
    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 15
    ws3.column_dimensions["E"].width = 10

    # Sheet 4: Analisis & Saran
    ws4 = wb.create_sheet("Analisis & Saran")
    headers_4 = ["Bulan", "Metrik", "Nilai", "Status", "Saran"]
    for col_idx, h in enumerate(headers_4, 1):
        ws4.cell(row=1, column=col_idx, value=h)
    style_header(ws4, len(headers_4))

    analytics_rows = []
    for m in range(1, 13):
        month_name = get_month_name(m)
        inc_val = sum(r.get("Amount", 0) or 0 for r in sdf_income.filter({"Month": month_name, "Year": year})) if sdf_income else 0.0
        exp_val = sum(r.get("Amount", 0) or 0 for r in sdf_expense.filter({"Month": month_name, "Year": year})) if sdf_expense else 0.0

        if inc_val > 0:
            ratio_80 = (exp_val / inc_val) * 100
            if ratio_80 >= 80:
                excess = exp_val - (0.8 * inc_val)
                analytics_rows.append({
                    "Bulan": month_name, "Metrik": "Rasio Pengeluaran vs Pemasukan",
                    "Nilai": f"{ratio_80:.1f}%", "Status": "⚠️ DARURAT",
                    "Saran": f"Potong pengeluaran minimal Rp{excess:,.0f} bulan depan"
                })

            exp_month_data = sdf_expense.filter({"Month": month_name, "Year": year})
            if exp_month_data:
                non_ess_total = sum(
                    r.get("Amount", 0) or 0 for r in exp_month_data
                    if any(kw in (r.get("Category", "") or "").lower()
                           for kw in ["hiburan", "gaya hidup", "game", "voucher", "streaming", "hobi"])
                )
                if non_ess_total > 0:
                    ne_ratio = (non_ess_total / inc_val) * 100
                    if ne_ratio > 20:
                        excess_ne = non_ess_total - (0.2 * inc_val)
                        analytics_rows.append({
                            "Bulan": month_name, "Metrik": "Rasio Non-Esensial",
                            "Nilai": f"{ne_ratio:.1f}%", "Status": "⚠️ PERHATIAN",
                            "Saran": f"Kurangi pengeluaran non-esensial Rp{excess_ne:,.0f}"
                        })

    if analytics_rows:
        for row_idx, r in enumerate(analytics_rows, 2):
            for col_idx, key in enumerate(headers_4, 1):
                style_data_cell(ws4.cell(row=row_idx, column=col_idx, value=r.get(key, "")))

    ws4.column_dimensions["A"].width = 15
    ws4.column_dimensions["B"].width = 35
    ws4.column_dimensions["C"].width = 15
    ws4.column_dimensions["D"].width = 18
    ws4.column_dimensions["E"].width = 50

    wb.save(output)
    wb.close()
    output.seek(0)
    return output


# =============================================================================
# RENDER: DELETE ITEMS SECTION (used by both income and expense)
# =============================================================================

def render_delete_section(data_type: str):
    """
    Render a delete section for either 'income' or 'expense' data.
    Shows all items in a table with delete buttons.
    """
    if data_type == "income":
        sdf = load_income_data()
        label = "Pemasukan"
        color = "#27AE60"
        columns_display = ["Tanggal", "Sumber", "Jumlah"]
    else:
        sdf = load_expense_data()
        label = "Pengeluaran"
        color = "#E74C3C"
        columns_display = ["Tanggal", "Kategori", "Jumlah"]

    if not sdf:
        st.info(f"Belum ada data {label.lower()}.")
        return

    st.markdown(f"#### Hapus Data {label}")

    # Display all items with row numbers and delete buttons
    all_data = sdf.to_dicts()

    for idx in range(len(all_data) - 1, -1, -1):
        r = all_data[idx]
        d = format_date_for_display(r.get("Date", ""))
        amount = format_rupiah(r.get("Amount", 0) or 0)

        if data_type == "income":
            desc = r.get("Source", "")
        else:
            desc = r.get("Category", "")

        col_a, col_b, col_c, col_d = st.columns([2, 3, 2, 1])
        with col_a:
            st.write(f"**{d}**")
        with col_b:
            st.write(desc)
        with col_c:
            st.write(amount)
        with col_d:
            if st.button(f"🗑️", key=f"del_{data_type}_{idx}"):
                # Remove and save
                sdf.remove_at(idx)
                if data_type == "income":
                    save_income_data(sdf)
                else:
                    save_expense_data(sdf)
                st.success(f"✅ Data {label} berhasil dihapus!")
                st.rerun()

    st.caption(f"Total {len(all_data)} item")


# =============================================================================
# UI COMPONENTS
# =============================================================================

def render_income_input():
    st.markdown("### 💰 Formulir Pemasukan")

    with st.container(border=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            input_date = st.date_input(
                "📅 Tanggal", value=st.session_state.income_date,
                key="income_date_input")
        with col2:
            input_source = st.text_input(
                "🏷️ Sumber Pemasukan", placeholder="Contoh: Gaji, Freelance, dll.",
                key="income_source_input")
        with col3:
            input_amount = st.number_input(
                "💵 Nominal (Rp)", min_value=0.0, step=10000.0, format="%.0f",
                key="income_amount_input")

        if input_date != st.session_state.income_date or \
           input_source != st.session_state.income_source or \
           input_amount != st.session_state.income_amount:
            st.session_state.income_submitted = False

        submitted = st.button("💾 Simpan Pemasukan", type="primary", use_container_width=True)

        if submitted:
            if input_amount <= 0:
                st.error("❌ Nominal pemasukan harus lebih dari 0!")
            elif not input_source.strip():
                st.error("❌ Sumber pemasukan tidak boleh kosong!")
            else:
                sdf_income = load_income_data()
                month_name = MONTH_NAMES_ID[input_date.month - 1]
                new_row = {
                    "Date": input_date, "Source": input_source.strip(),
                    "Amount": input_amount, "Month": month_name, "Year": input_date.year
                }
                sdf_income.append(new_row)
                save_income_data(sdf_income)
                st.session_state.income_submitted = True
                st.session_state.income_date = date.today()
                st.session_state.income_source = ""
                st.session_state.income_amount = 0.0
                st.success(f"✅ Pemasukan **{format_rupiah(input_amount)}** dari **{input_source.strip()}** berhasil disimpan!")
                st.balloons()
                st.rerun()

    sdf_income = load_income_data()
    if sdf_income:
        total_income = sdf_income.sum_column("Amount")
        st.info(f"📊 **Total Pemasukan Tersimpan:** {format_rupiah(total_income)}")


def render_expense_input():
    st.markdown("### 🛒 Formulir Pengeluaran")

    with st.container(border=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            input_date = st.date_input(
                "📅 Tanggal", value=st.session_state.expense_date,
                key="expense_date_input")
        with col2:
            input_category = st.text_input(
                "📂 Kategori Pengeluaran",
                placeholder="Contoh: Makanan, Transportasi, Tagihan, Hiburan, Gaya Hidup, dll.",
                key="expense_category_input")
        with col3:
            input_amount = st.number_input(
                "💵 Nominal (Rp)", min_value=0.0, step=5000.0, format="%.0f",
                key="expense_amount_input")

        if input_date != st.session_state.expense_date or \
           input_category != st.session_state.expense_category or \
           input_amount != st.session_state.expense_amount:
            st.session_state.expense_submitted = False

        submitted = st.button("💾 Simpan Pengeluaran", type="primary", use_container_width=True)

        if submitted:
            if input_amount <= 0:
                st.error("❌ Nominal pengeluaran harus lebih dari 0!")
            elif not input_category.strip():
                st.error("❌ Kategori pengeluaran tidak boleh kosong!")
            else:
                sdf_expense = load_expense_data()
                month_name = MONTH_NAMES_ID[input_date.month - 1]
                new_row = {
                    "Date": input_date, "Category": input_category.strip(),
                    "Amount": input_amount, "Month": month_name, "Year": input_date.year
                }
                sdf_expense.append(new_row)
                save_expense_data(sdf_expense)
                st.session_state.expense_submitted = True
                st.session_state.expense_date = date.today()
                st.session_state.expense_category = ""
                st.session_state.expense_amount = 0.0
                st.success(f"✅ Pengeluaran **{format_rupiah(input_amount)}** untuk **{input_category.strip()}** berhasil disimpan!")
                st.snow()
                st.rerun()

    sdf_expense = load_expense_data()
    if sdf_expense:
        total_expense = sdf_expense.sum_column("Amount")
        st.info(f"📊 **Total Pengeluaran Tersimpan:** {format_rupiah(total_expense)}")


def render_dashboard():
    st.markdown("## 📊 Dashboard Keuangan")

    sdf_income = load_income_data()
    sdf_expense = load_expense_data()

    col_year, col_month, _ = st.columns([1, 1, 2])
    with col_year:
        available_years = []
        if sdf_income:
            available_years.extend(sdf_income.unique("Year"))
        if sdf_expense:
            available_years.extend(sdf_expense.unique("Year"))
        available_years = sorted(set(available_years), reverse=True)
        if not available_years:
            available_years = [datetime.now().year]
        selected_year = st.selectbox("📆 Tahun", options=available_years, index=0)

    with col_month:
        selected_month = st.selectbox("🗓️ Bulan", options=list(range(1, 13)),
                                       format_func=lambda x: MONTH_NAMES_ID[x - 1])

    total_income, total_expense, net_income, savings_rate, inc_month, exp_month = calculate_kpis(
        sdf_income, sdf_expense, selected_month, selected_year
    )
    month_name = MONTH_NAMES_ID[selected_month - 1]

    st.markdown(f"### Ringkasan — {month_name} {selected_year}")
    kpi_cols = st.columns(4)

    with kpi_cols[0]:
        st.metric(label="💰 Total Pemasukan", value=format_rupiah(total_income))
    with kpi_cols[1]:
        st.metric(label="💸 Total Pengeluaran", value=format_rupiah(total_expense), delta_color="inverse")
    with kpi_cols[2]:
        delta_color = "normal" if net_income >= 0 else "inverse"
        st.metric(label="💵 Saldo Bersih", value=format_rupiah(net_income),
                  delta=f"{'✅ Surplus' if net_income >= 0 else '❌ Defisit'}", delta_color=delta_color)
    with kpi_cols[3]:
        if savings_rate is not None:
            st.metric(label="🏦 Rasio Tabungan", value=f"{savings_rate:.1f}%",
                      delta=f"{'✅ Sehat' if savings_rate >= 20 else '⚠️ Perlu Ditingkatkan'}")
        else:
            st.metric(label="🏦 Rasio Tabungan", value="N/A", delta="Belum ada pemasukan")

    st.markdown("### 📈 Perbandingan Pemasukan vs Pengeluaran Bulanan")
    if sdf_income or sdf_expense:
        fig = create_comparison_chart(sdf_income, sdf_expense, selected_year)
        if fig is not None:
            st.plotly_chart(fig, use_container_width=True, config={
                "displayModeBar": True, "displaylogo": False,
                "modeBarButtonsToRemove": ["lasso2d", "select2d"]})
    else:
        st.info("📭 Belum ada data keuangan. Silakan masukkan pemasukan atau pengeluaran terlebih dahulu.")

    st.markdown("### 📋 Detail Transaksi Bulanan")
    tab1, tab2 = st.tabs(["📈 Pemasukan", "📉 Pengeluaran"])

    with tab1:
        if inc_month:
            table_data = []
            for r in inc_month:
                table_data.append({
                    "Tanggal": format_date_for_display(r.get("Date", "")),
                    "Sumber": r.get("Source", ""),
                    "Jumlah": format_rupiah(r.get("Amount", 0) or 0)
                })
            st.dataframe(table_data, use_container_width=True, hide_index=True,
                         column_config={"Tanggal": "Tanggal", "Sumber": "Sumber", "Jumlah": "Jumlah"})
            st.caption(f"Total: {format_rupiah(total_income)} | {len(inc_month)} transaksi")
        else:
            st.info(f"📭 Belum ada data pemasukan untuk {month_name} {selected_year}.")

    with tab2:
        if exp_month:
            table_data = []
            for r in exp_month:
                table_data.append({
                    "Tanggal": format_date_for_display(r.get("Date", "")),
                    "Kategori": r.get("Category", ""),
                    "Jumlah": format_rupiah(r.get("Amount", 0) or 0)
                })
            cat_summary = {}
            for r in exp_month:
                cat = r.get("Category", "")
                amt = r.get("Amount", 0) or 0
                cat_summary[cat] = cat_summary.get(cat, 0) + amt
            cat_table = [{"Kategori": k, "Jumlah": format_rupiah(v)} for k, v in sorted(cat_summary.items())]

            col_tb1, col_tb2 = st.columns([2, 1])
            with col_tb1:
                st.dataframe(table_data, use_container_width=True, hide_index=True,
                             column_config={"Tanggal": "Tanggal", "Kategori": "Kategori", "Jumlah": "Jumlah"})
                st.caption(f"Total: {format_rupiah(total_expense)} | {len(exp_month)} transaksi")
            with col_tb2:
                st.markdown("**Rincian per Kategori**")
                st.dataframe(cat_table, use_container_width=True, hide_index=True)
        else:
            st.info(f"📭 Belum ada data pengeluaran untuk {month_name} {selected_year}.")


def render_analytics():
    st.markdown("## 🧠 Perspective Analytics")
    st.caption("Evaluasi pola pengeluaran berdasarkan aturan keuangan sehat (50/30/20)")

    sdf_income = load_income_data()
    sdf_expense = load_expense_data()

    if not sdf_income and not sdf_expense:
        st.info("📭 Belum ada data untuk dianalisis. Silakan masukkan data terlebih dahulu.")
        return

    col_year, col_month, _ = st.columns([1, 1, 2])
    with col_year:
        available_years = []
        if sdf_income:
            available_years.extend(sdf_income.unique("Year"))
        if sdf_expense:
            available_years.extend(sdf_expense.unique("Year"))
        available_years = sorted(set(available_years), reverse=True)
        if not available_years:
            available_years = [datetime.now().year]
        selected_year = st.selectbox("📆 Tahun", options=available_years, index=0, key="analytics_year")

    with col_month:
        selected_month = st.selectbox("🗓️ Bulan", options=list(range(1, 13)),
                                       format_func=lambda x: MONTH_NAMES_ID[x - 1],
                                       key="analytics_month")

    month_name = MONTH_NAMES_ID[selected_month - 1]
    st.markdown(f"### Analisis untuk {month_name} {selected_year}")

    results = run_perspective_analytics(sdf_income, sdf_expense, selected_month, selected_year)

    if not results:
        st.success(f"✅ Pola keuangan Anda untuk {month_name} {selected_year} dalam kondisi sehat! Tidak ada peringatan.")
        return

    for result in results:
        if result["type"] == "error":
            with st.container(border=True):
                st.error(result["title"])
                st.markdown(result["message"])
                if result.get("suggestion"):
                    st.warning(result["suggestion"])
        elif result["type"] == "warning":
            with st.container(border=True):
                st.warning(result["title"])
                st.markdown(result["message"])
                if result.get("suggestion"):
                    st.info(result["suggestion"])
        else:
            with st.container(border=True):
                st.info(result["title"])
                st.markdown(result["message"])

    st.markdown("### 📊 Statistik Tambahan")
    total_income, total_expense, net_income, savings_rate, _, _ = calculate_kpis(
        sdf_income, sdf_expense, selected_month, selected_year
    )

    col_s1, col_s2, col_s3, col_s4 = st.columns(4)
    with col_s1:
        if total_income > 0:
            st.metric("🏠 Ideal Kebutuhan (50%)", format_rupiah(total_income * 0.5))
        else:
            st.metric("🏠 Ideal Kebutuhan (50%)", "N/A")
    with col_s2:
        if total_income > 0:
            st.metric("🎮 Ideal Keinginan (30%)", format_rupiah(total_income * 0.3))
        else:
            st.metric("🎮 Ideal Keinginan (30%)", "N/A")
    with col_s3:
        if total_income > 0:
            st.metric("💰 Ideal Tabungan (20%)", format_rupiah(total_income * 0.2))
        else:
            st.metric("💰 Ideal Tabungan (20%)", "N/A")
    with col_s4:
        if savings_rate is not None:
            st.metric("🏦 Realisasi Tabungan", f"{savings_rate:.1f}%")
        else:
            st.metric("🏦 Realisasi Tabungan", "N/A")

    if PLOTLY_AVAILABLE and sdf_expense:
        exp_month_data = sdf_expense.filter({"Month": month_name, "Year": selected_year})
        if exp_month_data:
            st.markdown("### 🥧 Distribusi Pengeluaran per Kategori")
            cat_totals = {}
            for r in exp_month_data:
                cat = r.get("Category", "")
                amt = r.get("Amount", 0) or 0
                cat_totals[cat] = cat_totals.get(cat, 0) + amt

            colors = {
                "Makanan": "#E74C3C", "Transportasi": "#3498DB",
                "Tagihan": "#F39C12", "Hiburan": "#9B59B6", "Gaya Hidup": "#1ABC9C"
            }

            fig_pie = px.pie(
                names=list(cat_totals.keys()), values=list(cat_totals.values()),
                title=f"Distribusi Pengeluaran — {month_name} {selected_year}",
                color=list(cat_totals.keys()), color_discrete_map=colors, hole=0.3
            )
            fig_pie.update_traces(textposition="inside", textinfo="percent+label",
                                   hovertemplate="<b>%{label}</b><br>Jumlah: Rp %{value:,.0f}<br>Persentase: %{percent}")
            fig_pie.update_layout(template="plotly_white", height=400, showlegend=True)
            st.plotly_chart(fig_pie, use_container_width=True, config={
                "displaylogo": False, "modeBarButtonsToRemove": ["lasso2d", "select2d"]})


def render_export():
    st.markdown("## 📥 Export Laporan Tahunan")

    sdf_income = load_income_data()
    sdf_expense = load_expense_data()

    if not sdf_income and not sdf_expense:
        st.info("📭 Belum ada data untuk diexport.")
        return

    col_year, _ = st.columns([1, 3])
    with col_year:
        available_years = []
        if sdf_income:
            available_years.extend(sdf_income.unique("Year"))
        if sdf_expense:
            available_years.extend(sdf_expense.unique("Year"))
        available_years = sorted(set(available_years), reverse=True)
        if not available_years:
            available_years = [datetime.now().year]
        export_year = st.selectbox("📆 Pilih Tahun untuk Export", options=available_years, index=0)

    st.markdown(f"""
    **Laporan tahunan {export_year} akan mencakup:**
    - 📊 Ringkasan keuangan per bulan
    - 📈 Detail seluruh transaksi pemasukan
    - 📉 Detail seluruh transaksi pengeluaran
    - 🧠 Analisis dan saran penghematan
    """)

    if st.button("📥 Generate & Download Laporan", type="primary", use_container_width=True):
        with st.spinner("Membuat laporan..."):
            try:
                report_buffer = generate_annual_report(sdf_income, sdf_expense, export_year)
                st.success(f"✅ Laporan tahunan {export_year} berhasil dibuat!")
                st.download_button(
                    label="⬇️ Klik untuk Download", data=report_buffer,
                    file_name=f"Laporan_Keuangan_{export_year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            except Exception as e:
                st.error(f"❌ Gagal membuat laporan: {str(e)}")


# =============================================================================
# MAIN APPLICATION
# =============================================================================

def main():
    st.set_page_config(
        page_title="Personal Finance Dashboard",
        page_icon="💰", layout="wide", initial_sidebar_state="expanded"
    )

    initialize_database()
    init_session_state()

    # Sidebar
    st.sidebar.markdown("# 💰 Finance Dashboard")
    st.sidebar.markdown("---")
    tab_options = ["📊 Dashboard", "✏️ Input Data", "🗑️ Hapus Data", "🧠 Analytics", "📥 Export"]
    selected_tab = st.sidebar.radio("Navigasi", options=tab_options, index=0, key="sidebar_nav")

    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📊 Status Database")
    sdf_income = load_income_data()
    sdf_expense = load_expense_data()
    income_count = len(sdf_income) if sdf_income else 0
    expense_count = len(sdf_expense) if sdf_expense else 0
    st.sidebar.metric("📈 Total Pemasukan", income_count)
    st.sidebar.metric("📉 Total Pengeluaran", expense_count)
    if sdf_income:
        st.sidebar.caption(f"💰 {format_rupiah(sdf_income.sum_column('Amount'))}")
    if sdf_expense:
        st.sidebar.caption(f"💸 {format_rupiah(sdf_expense.sum_column('Amount'))}")
    st.sidebar.markdown("---")
    st.sidebar.caption(f"🕒 {datetime.now().strftime('%d-%m-%Y %H:%M')}")

    # Main content
    st.title("💰 Personal Finance Dashboard")
    st.markdown("Sistem pencatatan, visualisasi, dan analisis keuangan pribadi berbasis data")
    st.divider()

    if selected_tab == "📊 Dashboard":
        render_dashboard()

    elif selected_tab == "✏️ Input Data":
        st.markdown("## ✏️ Input Data Keuangan")
        st.markdown("Masukkan data pemasukan dan pengeluaran Anda secara dinamis.")
        tab_input1, tab_input2 = st.tabs(["💰 Pemasukan", "🛒 Pengeluaran"])
        with tab_input1:
            render_income_input()
        with tab_input2:
            render_expense_input()

    elif selected_tab == "🗑️ Hapus Data":
        st.markdown("## 🗑️ Hapus Data")
        st.markdown("Hapus item pemasukan atau pengeluaran yang tidak diinginkan.")
        tab_del1, tab_del2 = st.tabs(["💰 Hapus Pemasukan", "🛒 Hapus Pengeluaran"])
        with tab_del1:
            render_delete_section("income")
        with tab_del2:
            render_delete_section("expense")

    elif selected_tab == "🧠 Analytics":
        render_analytics()

    elif selected_tab == "📥 Export":
        render_export()


if __name__ == "__main__":
    main()